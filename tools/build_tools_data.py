#!/usr/bin/env python3
"""AIツール一覧v2.xlsx から assets/js/tools-data.js を生成する。"""

from __future__ import annotations

import argparse
import json
import re
import shutil
import urllib.parse
import urllib.request
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[1]
TOOLS_PAGES_DIR = ROOT / "tools"
XLSX_DEFAULT = Path.home() / "Downloads" / "AIツール一覧v2.xlsx"
XLSX_PROJECT = ROOT / "data" / "ai-tools-v2.xlsx"
OUT_JS = ROOT / "assets" / "js" / "tools-data.js"
ICON_DIR = ROOT / "assets" / "images" / "tools"

CAT_MAP = {
    "チャットAI": "chat",
    "AIエージェント": "agent",
    "画像生成": "image",
    "コーディング": "code",
    "音声・動画・音楽": "audio",
    "仕事効率化・資料・翻訳": "productivity",
    "検索・リサーチ": "research",
    "業務自動化": "automation",
    "ローカル・上級者向け": "local",
}

FILTER_LABELS = {
    "chat": "チャットAI",
    "agent": "AIエージェント",
    "image": "画像生成",
    "code": "コーディング",
    "audio": "音声・動画",
    "productivity": "仕事効率化",
    "research": "検索・リサーチ",
    "automation": "業務自動化",
    "local": "ローカル",
}

FEATURED_IDS = {"chatgpt", "claude", "gemini"}
MIN_RASTER_BYTES = 2500
USER_AGENT = (
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
    "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/122.0.0.0 Safari/537.36"
)

# メーカー共通 SVG（個別アイコンが取れないときのフォールバック）
SHARED_LOGO = {
    "chatgpt": "openai.svg",
    "dalle": "openai.svg",
    "whisper": "openai.svg",
    "claude": "anthropic.svg",
    "gemini": "google.svg",
    "notebooklm": "google.svg",
    "veo-3": "google.svg",
    "copilot": "microsoft.svg",
    "bing-image-creator": "microsoft.svg",
    "github-copilot": "github.svg",
    "adobe-firefly": "adobe.svg",
    "adobe-podcast": "adobe.svg",
    "adobe-acrobat-ai": "adobe.svg",
    "canva": "canva.svg",
    "deepl": "deepl.svg",
    "notion-ai": "notion.svg",
    "perplexity": "perplexity.svg",
    "replit": "replit.svg",
    "v0": "vercel.svg",
    "bolt-new": "bolt-new.svg",
    "chatgpt-operator": "openai.svg",
    "codex-cli": "openai.svg",
    "sora-2": "openai.svg",
    "claude-code": "anthropic.svg",
    "microsoft-365-copilot": "microsoft.svg",
    "microsoft-designer": "microsoft.svg",
    "microsoft-power-automate": "microsoft.svg",
    "google-workspace-ai": "google.svg",
    "adobe-express": "adobe.svg",
    "bing-image-creator": "microsoft.svg",
}

# 取得元ドメインの上書き（公式URLとアイコンホストが異なる場合）
ICON_DOMAIN_OVERRIDE = {
    "grok": "x.ai",
    "deepseek-chat": "www.deepseek.com",
    "le-chat": "mistral.ai",
    "stable-diffusion": "stability.ai",
    "amazon-q-developer": "aws.amazon.com",
    "veo-3": "deepmind.google",
    "otter-ai": "otter.ai",
    "tl-dv": "tldv.io",
    "pi": "pi.ai",
    "devin": "devin.ai",
}

# 直接ダウンロードURL（Bot対策で favicon が HTML になる場合）
DIRECT_ICON_URL = {
    "devin": "https://devin.ai/icon.png",
    "ollama": "https://ollama.com/public/apple-touch-icon.png",
    "bolt-new": "https://cdn.simpleicons.org/stackblitz/1389FD",
    # 記事用ヒーローアイコン（各社公式）
    "claude": "https://claude.ai/apple-touch-icon.png",
    "gemini": "https://www.gstatic.com/lamda/images/gemini_sparkle_v002_d4735304ff6292a690345.svg",
    "copilot": "https://copilot.microsoft.com/favicon.ico",
    "grok": "https://grok.com/apple-touch-icon.png",
    "meta-ai": "https://static.xx.fbcdn.net/assets/?set=meta_ai_assets&name=meta-ai-orbit-favicon-180px-raster&density=4x&mode=light",
    "poe": "https://psc2.cf2.poecdn.net/assets/apple-touch-icon.png",
    "deepseek-chat": "https://cdn.deepseek.com/chat/icon.png",
    "le-chat": "https://chat.mistral.ai/favicons/apple-touch-icon.png",
    "pi": "https://framerusercontent.com/images/Hu7aeJCxpUvwSxyA5mfRMSPAqAU.svg",
    "midjourney": "https://www.midjourney.com/public/apple-touch-icon.png",
    "dalle": "https://chatgpt.com/cdn/assets/favicon-32x32.png",
    "adobe-firefly": "https://www.adobe.com/favicon.ico",
    "stable-diffusion": "https://stability.ai/favicon.ico",
    "flux": "https://blackforestlabs.ai/favicon.ico",
    "ideogram": "https://ideogram.ai/favicon.ico",
    "canva": "https://www.canva.com/favicon.ico",
    "bing-image-creator": "https://www.bing.com/sa/simg/favicon-2x.ico",
    "github-copilot": "https://github.com/apple-touch-icon.png",
    "cursor": "https://cursor.com/apple-touch-icon.png",
    "windsurf": "https://www.google.com/s2/favicons?domain=codeium.com&sz=256",
    "replit": "https://replit.com/public/images/replit-logo-small.png",
    "v0": "https://www.google.com/s2/favicons?domain=v0.dev&sz=256",
    "bolt-new": "https://cdn.simpleicons.org/stackblitz/1389FD",
    "lovable": "https://lovable.dev/apple-touch-icon.png",
    "tabnine": "https://www.tabnine.com/wp-content/uploads/2024/09/cropped-tabnine-favicon-270x270.png",
    "amazon-q-developer": "https://d1.awsstatic.com/onedam/marketing-channels/website/aws/en_US/product-categories/business-application/approved/images/qdev_merch_qdev_v1_1280x1280.e0fc098598c9a070a7a5c866716c387fba743c1e.jpg",
    "warp": "https://www.warp.dev/android-chrome-512x512.png",
    "elevenlabs": "https://elevenlabs.io/apple-icon.png",
    "runway": "https://runwayml.com/icon.png",
    "kling-ai": "https://s16-kling.klingai.com/kos/s101/nlav112918/kling-homepage-aio/logo-180x180.png",
    "pika": "https://pika.art/images/landing/footer-pika-icon.webp",
    "luma-dream-machine": "https://lumalabs.ai/images/brand/luma-ai/logo-black.svg",
    "heygen": "https://cdn.sanity.io/images/pdhqcmb1/production/9326953a01da96181ed910f1ad68214477fe9846-264x264.png",
    "suno": "https://cdn-o.suno.com/apple-touch-icon.png",
    "udio": "https://www.udio.com/favicon.ico?favicon.0wbc3vz33genq.ico",
    "descript": "https://static-cdn.descript.com/web/icons/apple-touch-icon.png",
    "capcut": "https://play-lh.googleusercontent.com/M78HyakHaxKrjoeqYx41E9DXfVYYtx67nvc7Ks4G4zFQeaAJdGCi8gzzGSrHIwlrmnJS6zD9S4fAXqdEwfuHQAQ=w512",
    "opus-clip": "https://cdn.prod.website-files.com/6388604483b03a9ecb34d695/6435197bfb1d6e486e04c37b_webclip.png",
}


NAME_TO_ID = {
    "イルシル": "irushiru",
}


def slugify(name: str) -> str:
    if name in NAME_TO_ID:
        return NAME_TO_ID[name]
    base = re.sub(r"（[^）]*）", "", name).strip()
    base = base.replace("·", "-")
    slug = re.sub(r"[^a-zA-Z0-9]+", "-", base).strip("-").lower()
    aliases = {
        "dall-e": "dalle",
        "microsoft-copilot": "copilot",
        "deepseek-chat": "deepseek-chat",
        "le-chat": "le-chat",
        "meta-ai": "meta-ai",
        "github-copilot": "github-copilot",
        "amazon-q-developer": "amazon-q-developer",
        "bolt-new": "bolt-new",
        "veo-3": "veo-3",
        "kling-ai": "kling-ai",
        "luma-dream-machine": "luma-dream-machine",
        "opus-clip": "opus-clip",
        "adobe-podcast": "adobe-podcast",
        "adobe-acrobat-ai": "adobe-acrobat-ai",
        "bing-image-creator": "bing-image-creator",
        "lm-studio": "lm-studio",
        "stable-diffusion": "stable-diffusion",
        "notion-ai": "notion-ai",
        "raycast-ai": "raycast-ai",
        "fireflies-ai": "fireflies-ai",
        "chatgpt-operator": "chatgpt-operator",
        "claude-code": "claude-code",
        "continue-dev": "continue-dev",
        "microsoft-365-copilot": "microsoft-365-copilot",
        "microsoft-designer": "microsoft-designer",
        "microsoft-power-automate": "microsoft-power-automate",
        "google-workspace-ai": "google-workspace-ai",
        "slack-ai": "slack-ai",
        "zoom-ai-companion": "zoom-ai-companion",
        "adobe-express": "adobe-express",
        "codex-cli": "codex-cli",
        "haiper-ai": "haiper-ai",
        "hailuo-ai": "hailuo-ai",
        "napkin-ai": "napkin-ai",
        "sora-2": "sora-2",
        "copy-ai": "copy-ai",
        "surfer-ai": "surfer-ai",
        "irushiru": "irushiru",
    }
    return aliases.get(slug, slug) or NAME_TO_ID.get(name, slug)


def load_rows(xlsx: Path) -> list[dict]:
    wb = openpyxl.load_workbook(xlsx, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = []
    for r in range(2, ws.max_row + 1):
        cat = ws.cell(r, 1).value
        name = ws.cell(r, 2).value
        maker = ws.cell(r, 3).value
        url = ws.cell(r, 4).value
        note = ws.cell(r, 5).value
        if not name or not cat:
            continue
        cat_name = str(cat).strip()
        cat_id = CAT_MAP[cat_name]
        tool_id = slugify(str(name).strip())
        rows.append(
            {
                "id": tool_id,
                "name": str(name).strip(),
                "maker": str(maker).strip() if maker else "",
                "url": str(url).strip() if url else "",
                "tagline": str(note).strip() if note else "",
                "cat": cat_id,
                "catLabel": cat_name,
                "featured": tool_id in FEATURED_IDS,
            }
        )
    return rows


def fetch_bytes(url: str) -> bytes | None:
    try:
        req = urllib.request.Request(url, headers={"User-Agent": USER_AGENT})
        with urllib.request.urlopen(req, timeout=20) as res:
            data = res.read()
        return data if data and len(data) >= 80 else None
    except Exception:
        return None


def is_image(data: bytes) -> bool:
    head = data[:256].lstrip().lower()
    if head.startswith((b"<!doctype", b"<html", b"<?xml version")) and b"<svg" not in head:
        return False
    if data.startswith(b"\x89PNG") or data.startswith(b"\xff\xd8\xff"):
        return True
    if data.startswith((b"GIF87a", b"GIF89a", b"RIFF")):
        return True
    if data[:4] in {b"\x00\x00\x01\x00", b"\x00\x00\x02\x00"}:
        return True
    return head.startswith((b"<svg", b"<?xml"))


def ext_for(data: bytes, default: str) -> str:
    if data.startswith(b"\x89PNG"):
        return ".png"
    if data.startswith(b"\xff\xd8\xff"):
        return ".jpg"
    if data[:4] in {b"\x00\x00\x01\x00", b"\x00\x00\x02\x00"}:
        return ".ico"
    head = data[:256].lstrip()
    if head.startswith((b"<svg", b"<?xml")):
        return ".svg"
    return default


def icon_hosts(tool_id: str, page_url: str) -> list[str]:
    parsed = urllib.parse.urlparse(page_url)
    hosts: list[str] = []
    if tool_id in ICON_DOMAIN_OVERRIDE:
        hosts.append(ICON_DOMAIN_OVERRIDE[tool_id])
    if parsed.netloc:
        hosts.append(parsed.netloc)
        bare = parsed.netloc.removeprefix("www.")
        if bare not in hosts:
            hosts.append(bare)
    return hosts


def icon_candidates(host: str) -> list[tuple[str, str]]:
    return [
        (f"https://icons.duckduckgo.com/ip3/{host}.ico", ".ico"),
        (f"https://www.google.com/s2/favicons?domain={host}&sz=128", ".png"),
        (f"https://{host}/apple-touch-icon.png", ".png"),
        (f"https://{host}/apple-touch-icon-precomposed.png", ".png"),
        (f"https://{host}/favicon-32x32.png", ".png"),
        (f"https://{host}/favicon.png", ".png"),
        (f"https://{host}/favicon.svg", ".svg"),
        (f"https://{host}/favicon.ico", ".ico"),
        (f"https://www.google.com/s2/favicons?domain={host}&sz=64", ".png"),
    ]


def download_best_icon(tool_id: str, page_url: str) -> str | None:
    best_data: bytes | None = None
    best_ext = ".png"
    best_score = 0

    if tool_id in DIRECT_ICON_URL:
        data = fetch_bytes(DIRECT_ICON_URL[tool_id])
        if data and is_image(data):
            best_data = data
            best_ext = ext_for(data, ".png")
            best_score = len(data) + (500 if best_ext == ".svg" else 0)

    for host in icon_hosts(tool_id, page_url):
        for url, default_ext in icon_candidates(host):
            data = fetch_bytes(url)
            if not data or not is_image(data):
                continue
            ext = ext_for(data, default_ext)
            score = len(data) + (500 if ext == ".svg" else 0)
            if score > best_score:
                best_data, best_ext, best_score = data, ext, score

    if not best_data:
        return None

    dest = ICON_DIR / f"{tool_id}{best_ext}"
    dest.write_bytes(best_data)
    return dest.name


def is_raster(filename: str) -> bool:
    return Path(filename).suffix.lower() in {".png", ".ico", ".jpg", ".jpeg", ".webp"}


def logo_file_valid(path: Path) -> bool:
    if not path.exists():
        return False
    data = path.read_bytes()
    return is_image(data)


def logo_quality_ok(path: Path) -> bool:
    if not logo_file_valid(path):
        return False
    size = path.stat().st_size
    if path.suffix.lower() == ".svg":
        return size >= 150
    return size >= MIN_RASTER_BYTES


def ensure_logo(tool: dict, force: bool = False) -> str:
    tool_id = tool["id"]
    page_url = tool["url"]

    if not force:
        for pattern in (f"{tool_id}.png", f"{tool_id}.ico", f"{tool_id}.svg", f"{tool_id}.webp"):
            candidate = ICON_DIR / pattern
            if logo_quality_ok(candidate):
                return candidate.name

        if tool_id in SHARED_LOGO:
            shared = ICON_DIR / SHARED_LOGO[tool_id]
            if logo_quality_ok(shared):
                return SHARED_LOGO[tool_id]

    fetched = None
    if page_url:
        fetched = download_best_icon(tool_id, page_url)
        if fetched and logo_quality_ok(ICON_DIR / fetched):
            return fetched

    if tool_id in SHARED_LOGO:
        shared = ICON_DIR / SHARED_LOGO[tool_id]
        if logo_file_valid(shared):
            return SHARED_LOGO[tool_id]

    if fetched and logo_file_valid(ICON_DIR / fetched):
        return fetched

    return "openai.svg"


def js_string(value: str) -> str:
    return json.dumps(value, ensure_ascii=False)


def has_article(tool_id: str) -> bool:
    return (TOOLS_PAGES_DIR / tool_id / "index.html").is_file()


def render_js(tools: list[dict]) -> str:
    lines = ["const TOOLS = ["]
    current_cat = None
    for tool in tools:
        if tool["catLabel"] != current_cat:
            current_cat = tool["catLabel"]
            lines.append(f"  // ── {current_cat} ──")
        lines.append("  {")
        lines.append(f"    id: {js_string(tool['id'])},")
        lines.append(f"    name: {js_string(tool['name'])},")
        lines.append(f"    maker: {js_string(tool['maker'])},")
        lines.append(f"    logo: {js_string(tool['logo'])},")
        lines.append(f"    cat: {js_string(tool['cat'])},")
        lines.append(f"    catLabel: {js_string(tool['catLabel'])},")
        if tool["featured"]:
            lines.append("    featured: true,")
        if has_article(tool["id"]):
            lines.append("    article: true,")
        lines.append(f"    tagline: {js_string(tool['tagline'])},")
        lines.append(f"    url: {js_string(tool['url'])},")
        lines.append("  },")
    lines.append("];")
    lines.append("")
    lines.append("const CATEGORIES = [")
    lines.append("  { id: 'all', label: 'すべて' },")
    for cat_id, label in FILTER_LABELS.items():
        lines.append(f"  {{ id: {js_string(cat_id)}, label: {js_string(label)} }},")
    lines.append("];")
    lines.append("")
    lines.append("const PAGE_SIZE = 20;")
    lines.append("")
    return "\n".join(lines)


def load_tools_from_js() -> list[dict]:
    text = OUT_JS.read_text(encoding="utf-8")
    tools = []
    for block in re.finditer(r"\{[^{}]+\}", text.split("const TOOLS = [")[1].split("];")[0]):
        chunk = block.group(0)

        def pick(key: str) -> str:
            match = re.search(rf'{key}:\s*"([^"]*)"', chunk)
            return match.group(1) if match else ""

        tool_id = pick("id")
        if not tool_id:
            continue
        tools.append(
            {
                "id": tool_id,
                "name": pick("name"),
                "maker": pick("maker"),
                "url": pick("url"),
                "tagline": pick("tagline"),
                "cat": pick("cat"),
                "catLabel": pick("catLabel"),
                "featured": "featured: true" in chunk,
                "logo": pick("logo"),
            }
        )
    return tools


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument(
        "--icons-only",
        action="store_true",
        help="既存 tools-data.js の URL からアイコンのみ再取得する",
    )
    parser.add_argument(
        "--force-icons",
        action="store_true",
        help="品質に関わらず全アイコンを再取得する",
    )
    args = parser.parse_args()

    ICON_DIR.mkdir(parents=True, exist_ok=True)

    if args.icons_only:
        tools = load_tools_from_js()
        if not tools:
            raise SystemExit("tools-data.js からツールを読み込めませんでした")
    else:
        XLSX_PROJECT.parent.mkdir(parents=True, exist_ok=True)
        if XLSX_PROJECT.exists():
            src = XLSX_PROJECT
        elif XLSX_DEFAULT.exists():
            shutil.copy2(XLSX_DEFAULT, XLSX_PROJECT)
            src = XLSX_PROJECT
        else:
            raise SystemExit(f"Excel not found: {XLSX_PROJECT}")
        tools = load_rows(src)

    refreshed = 0
    for tool in tools:
        before = tool.get("logo", "")
        tool["logo"] = ensure_logo(tool, force=args.force_icons)
        if tool["logo"] != before:
            refreshed += 1

    OUT_JS.write_text(render_js(tools), encoding="utf-8")
    print(f"Wrote {len(tools)} tools -> {OUT_JS}")
    print(f"Icons updated: {refreshed}")
    if not args.icons_only:
        print(f"Source: {XLSX_PROJECT}")


if __name__ == "__main__":
    main()
