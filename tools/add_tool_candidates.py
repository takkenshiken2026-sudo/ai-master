#!/usr/bin/env python3
"""追加候補AIツール_v2.xlsx をマスターに取り込み、記事・画像・sitemap を生成する。"""

from __future__ import annotations

import html
import json
import re
import shutil
import urllib.parse
import urllib.request
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[1]
CANDIDATES_XLSX = Path.home() / "Downloads" / "追加候補AIツール_v2.xlsx"
MASTER_XLSX = ROOT / "data" / "ai-tools-v2.xlsx"
ICON_DIR = ROOT / "assets" / "images" / "tools"
SITEMAP = ROOT / "sitemap.xml"
USER_AGENT = (
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
    "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/122.0.0.0 Safari/537.36"
)

CAT_ORDER = [
    "チャットAI",
    "AIエージェント",
    "画像生成",
    "コーディング",
    "音声・動画・音楽",
    "仕事効率化・資料・翻訳",
    "検索・リサーチ",
    "業務自動化",
    "ローカル・上級者向け",
]

CAT_SHORT = {
    "チャットAI": "チャットAI",
    "AIエージェント": "AIエージェント",
    "画像生成": "画像生成",
    "コーディング": "コーディング",
    "音声・動画・音楽": "音声・動画",
    "仕事効率化・資料・翻訳": "仕事効率化",
    "検索・リサーチ": "検索・リサーチ",
    "業務自動化": "業務自動化",
    "ローカル・上級者向け": "ローカル",
}

NAME_TO_ID = {
    "イルシル": "irushiru",
}

TOOL_URLS: dict[str, str] = {
    "Manus": "https://manus.im",
    "ChatGPT Operator": "https://openai.com/index/introducing-operator/",
    "Skywork": "https://skywork.ai",
    "Claude Code": "https://docs.anthropic.com/en/docs/claude-code/overview",
    "Cline": "https://cline.bot",
    "Zapier": "https://zapier.com",
    "Recraft": "https://www.recraft.ai",
    "Haiper AI": "https://haiper.ai",
    "Adobe Express": "https://www.adobe.com/express/",
    "Microsoft Designer": "https://designer.microsoft.com",
    "Microsoft 365 Copilot": "https://www.microsoft.com/microsoft-365/copilot",
    "Slack AI": "https://slack.com/features/ai",
    "Zoom AI Companion": "https://www.zoom.com/en/products/ai-assistant/",
    "Google Workspace AI": "https://workspace.google.com/solutions/ai/",
    "イルシル": "https://irushiru.com",
    "Hailuo AI（MiniMax）": "https://hailuoai.com/video",
    "CoeFont": "https://coefont.cloud",
    "Make（旧Integromat）": "https://www.make.com",
    "n8n": "https://n8n.io",
    "Writesonic": "https://writesonic.com",
    "Jasper": "https://www.jasper.ai",
    "Codex CLI": "https://developers.openai.com/codex/cli",
    "Continue.dev": "https://www.continue.dev",
    "AutoGPT": "https://agpt.co",
    "Napkin AI": "https://www.napkin.ai",
    "Synthesia": "https://www.synthesia.io",
    "Dify": "https://dify.ai",
    "NoLang": "https://nolang.ai",
    "Sora 2": "https://openai.com/sora/",
    "Copy.ai": "https://www.copy.ai",
    "Surfer AI": "https://surferseo.com/surfer-ai/",
    "Microsoft Power Automate": "https://www.microsoft.com/power-platform/products/power-automate",
    "Magnific（旧Freepik AI）": "https://magnific.ai",
}

COMPARE_WITH: dict[str, list[str]] = {
    "manus": ["devin", "chatgpt-operator", "autogpt"],
    "chatgpt-operator": ["manus", "devin", "chatgpt"],
    "skywork": ["perplexity", "genspark", "gamma"],
    "claude-code": ["github-copilot", "cursor", "codex-cli"],
    "cline": ["cursor", "windsurf", "continue-dev"],
    "zapier": ["make", "n8n", "microsoft-power-automate"],
    "recraft": ["ideogram", "adobe-firefly", "canva"],
    "haiper-ai": ["runway", "pika", "veo-3"],
    "adobe-express": ["canva", "adobe-firefly", "microsoft-designer"],
    "microsoft-designer": ["bing-image-creator", "canva", "adobe-express"],
    "microsoft-365-copilot": ["notion-ai", "copilot", "google-workspace-ai"],
    "slack-ai": ["notion-ai", "microsoft-365-copilot", "fireflies-ai"],
    "zoom-ai-companion": ["otter-ai", "notta", "fireflies-ai"],
    "google-workspace-ai": ["gemini", "microsoft-365-copilot", "notebooklm"],
    "irushiru": ["gamma", "beautiful-ai", "canva"],
    "hailuo-ai": ["kling-ai", "pika", "runway"],
    "coefont": ["elevenlabs", "heygen", "notta"],
    "make": ["zapier", "n8n", "microsoft-power-automate"],
    "n8n": ["zapier", "make", "dify"],
    "writesonic": ["jasper", "copy-ai", "chatgpt"],
    "jasper": ["writesonic", "copy-ai", "chatgpt"],
    "codex-cli": ["claude-code", "github-copilot", "warp"],
    "continue-dev": ["github-copilot", "cline", "tabnine"],
    "autogpt": ["manus", "devin", "chatgpt-operator"],
    "napkin-ai": ["gamma", "beautiful-ai", "canva"],
    "synthesia": ["heygen", "runway", "elevenlabs"],
    "dify": ["ollama", "lm-studio", "n8n"],
    "nolang": ["pika", "capcut", "runway"],
    "sora-2": ["veo-3", "runway", "kling-ai"],
    "copy-ai": ["jasper", "writesonic", "chatgpt"],
    "surfer-ai": ["writesonic", "jasper", "chatgpt"],
    "microsoft-power-automate": ["zapier", "make", "microsoft-365-copilot"],
    "magnific": ["recraft", "leonardo-ai", "adobe-firefly"],
}

SHARED_LOGO_FOR_RELATED = {
    "chatgpt": "openai.svg",
    "claude": "anthropic.svg",
    "gemini": "google.svg",
    "copilot": "microsoft.svg",
    "dalle": "openai.svg",
    "whisper": "openai.svg",
    "github-copilot": "github.svg",
    "adobe-firefly": "adobe.svg",
    "canva": "canva.svg",
    "perplexity": "perplexity.svg",
    "notion-ai": "notion.svg",
    "deepl": "deepl.svg",
    "replit": "replit.svg",
}


def slugify(name: str) -> str:
    if name in NAME_TO_ID:
        return NAME_TO_ID[name]
    base = re.sub(r"（[^）]*）", "", name).strip()
    base = base.replace("·", "-")
    slug = re.sub(r"[^a-zA-Z0-9]+", "-", base).strip("-").lower()
    return slug or NAME_TO_ID.get(name, re.sub(r"[^a-z0-9]+", "-", name.lower()))


def short_tagline(reason: str, max_len: int = 52) -> str:
    text = str(reason or "").strip()
    if not text:
        return ""
    first = text.split("。")[0].strip()
    if len(first) <= max_len:
        return first
    return first[: max_len - 1] + "…"


def load_candidates() -> list[dict]:
    wb = openpyxl.load_workbook(CANDIDATES_XLSX, read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))[1:]
    items = []
    for row in rows:
        _prio, name, cat, maker, reason, title = row[:6]
        name = str(name).strip()
        items.append(
            {
                "name": name,
                "cat": str(cat).strip(),
                "maker": str(maker).strip(),
                "reason": str(reason or "").strip(),
                "title": str(title or f"{name}とは？機能・料金・使い方を解説").strip(),
                "url": TOOL_URLS.get(name, ""),
                "id": slugify(name),
                "tagline": short_tagline(str(reason or "")),
            }
        )
    return items


def merge_excel(candidates: list[dict]) -> None:
    wb = openpyxl.load_workbook(MASTER_XLSX)
    ws = wb.active
    existing = []
    for r in range(2, ws.max_row + 1):
        existing.append(
            (
                ws.cell(r, 1).value,
                ws.cell(r, 2).value,
                ws.cell(r, 3).value,
                ws.cell(r, 4).value,
                ws.cell(r, 5).value,
            )
        )
    existing_names = {str(x[1]).strip() for x in existing if x[1]}
    added = 0
    for c in candidates:
        if c["name"] in existing_names:
            continue
        existing.append((c["cat"], c["name"], c["maker"], c["url"], c["tagline"]))
        existing_names.add(c["name"])
        added += 1

    def sort_key(row):
        cat = str(row[0])
        try:
            cat_idx = CAT_ORDER.index(cat)
        except ValueError:
            cat_idx = 99
        return (cat_idx, str(row[1]))

    existing.sort(key=sort_key)
    if ws.max_row >= 2:
        ws.delete_rows(2, ws.max_row - 1)
    for i, row in enumerate(existing, start=2):
        for col, val in enumerate(row, start=1):
            ws.cell(i, col, val)
    wb.save(MASTER_XLSX)
    print(f"Excel: {added} 行追加 → 合計 {len(existing)} 行")


def fetch_bytes(url: str) -> bytes | None:
    try:
        req = urllib.request.Request(url, headers={"User-Agent": USER_AGENT})
        with urllib.request.urlopen(req, timeout=25) as res:
            data = res.read()
        return data if data and len(data) > 80 else None
    except Exception:
        return None


def find_og_image(page_url: str) -> str | None:
    data = fetch_bytes(page_url)
    if not data:
        return None
    text = data.decode("utf-8", errors="ignore")
    for pattern in (
        r'<meta[^>]+property=["\']og:image["\'][^>]+content=["\']([^"\']+)',
        r'<meta[^>]+content=["\']([^"\']+)["\'][^>]+property=["\']og:image',
        r'<meta[^>]+name=["\']twitter:image["\'][^>]+content=["\']([^"\']+)',
    ):
        m = re.search(pattern, text, re.I)
        if m:
            return urllib.parse.urljoin(page_url, m.group(1))
    return None


def save_image(data: bytes, dest: Path) -> None:
    dest.parent.mkdir(parents=True, exist_ok=True)
    dest.write_bytes(data)


def ensure_article_images(tool_id: str, page_url: str, maker: str) -> tuple[str, str]:
    """Returns (app_icon_rel, og_hero_rel) filenames under assets/images/tools/{id}/"""
    tool_dir = ICON_DIR / tool_id
    tool_dir.mkdir(parents=True, exist_ok=True)
    app_path = tool_dir / "app-icon.png"
    og_path = tool_dir / "og-hero.png"

    # Try list-level icon first
    for ext in (".png", ".ico", ".svg", ".webp"):
        src = ICON_DIR / f"{tool_id}{ext}"
        if src.is_file() and src.stat().st_size > 500:
            if ext != ".png":
                data = src.read_bytes()
            else:
                data = src.read_bytes()
            if ext == ".svg":
                app_path = tool_dir / "app-icon.svg"
            save_image(data, app_path)
            break
    else:
        hosts = []
        parsed = urllib.parse.urlparse(page_url)
        if parsed.netloc:
            hosts.append(parsed.netloc)
        for host in hosts:
            for url in (
                f"https://{host}/apple-touch-icon.png",
                f"https://www.google.com/s2/favicons?domain={host}&sz=256",
            ):
                data = fetch_bytes(url)
                if data:
                    save_image(data, app_path)
                    break
            if app_path.is_file():
                break

    og_url = find_og_image(page_url) if page_url else None
    if og_url:
        data = fetch_bytes(og_url)
        if data:
            ext = ".jpg" if data[:3] == b"\xff\xd8\xff" else ".png"
            og_path = tool_dir / f"og-hero{ext}"
            save_image(data, og_path)

    if not og_path.is_file() and app_path.is_file():
        try:
            from PIL import Image

            with Image.open(app_path) as img:
                img = img.convert("RGBA")
                w, h = img.size
                scale = max(800 / w, 512 / h, 1.0)
                nw, nh = int(w * scale), int(h * scale)
                img = img.resize((nw, nh), Image.Resampling.LANCZOS)
                canvas = Image.new("RGBA", (800, 450), (245, 247, 250, 255))
                canvas.paste(img, ((800 - nw) // 2, (450 - nh) // 2), img if img.mode == "RGBA" else None)
                canvas.convert("RGB").save(og_path, "PNG")
        except Exception:
            og_path = tool_dir / "og-hero.png"
            if app_path.is_file():
                shutil.copy2(app_path, og_path)

    app_name = app_path.name if app_path.is_file() else "app-icon.png"
    og_name = og_path.name if og_path.is_file() else "og-hero.png"
    return app_name, og_name


def faq_item(q: str, a: str) -> str:
    return (
        f'        <details>\n'
        f'          <summary>{html.escape(q)}</summary>\n'
        f'          <p class="tool-faq-a">{html.escape(a)}</p>\n'
        f'        </details>'
    )


def faq_json(q: str, a: str) -> str:
    return json.dumps({"@type": "Question", "name": q, "acceptedAnswer": {"@type": "Answer", "text": a}}, ensure_ascii=False)


def related_link(tool_id: str, name: str) -> str:
    logo = SHARED_LOGO_FOR_RELATED.get(tool_id)
    if logo:
        icon = f'<img src="../../assets/images/tools/{logo}" alt="" width="16" height="16"> '
    elif (ICON_DIR / tool_id / "app-icon.png").is_file():
        icon = f'<img src="../../assets/images/tools/{tool_id}/app-icon.png" alt="" width="16" height="16"> '
    elif (ICON_DIR / f"{tool_id}.png").is_file():
        icon = f'<img src="../../assets/images/tools/{tool_id}.png" alt="" width="16" height="16"> '
    else:
        icon = ""
    return f'          <li><a href="../{tool_id}/">{icon}{html.escape(name)}</a></li>'


def load_all_tool_names() -> dict[str, str]:
    text = (ROOT / "assets" / "js" / "tools-data.js").read_text(encoding="utf-8")
    names = {}
    for m in re.finditer(r'id:\s*"([^"]+)"[^}]*?name:\s*"([^"]+)"', text, re.S):
        names[m.group(1)] = m.group(2)
    return names


def generate_article(tool: dict, all_names: dict[str, str]) -> str:
    tid = tool["id"]
    name = tool["name"]
    maker = tool["maker"]
    cat_short = CAT_SHORT.get(tool["cat"], tool["cat"])
    title = tool["title"]
    if not title.endswith(" — AI Master"):
        page_title = f"{title} — AI Master"
    else:
        page_title = title
    headline = title.replace(" — AI Master", "")
    url = tool["url"]
    canonical = f"https://ai-master.jp/tools/{tid}/"
    reason = tool["reason"]
    compare_ids = [x for x in COMPARE_WITH.get(tid, []) if x in all_names][:3]
    if len(compare_ids) < 2:
        compare_ids = ["chatgpt", "claude", "gemini"]
    compare_names = [all_names.get(x, x) for x in compare_ids]

    app_icon, og_hero = ensure_article_images(tid, url, maker)
    og_url = f"https://ai-master.jp/assets/images/tools/{tid}/{og_hero}"

    misconception = (
        f"代表的な誤解は「<strong>{html.escape(name)}＝生成AIそのもの</strong>」と覚えることです。"
        f"{html.escape(name)}は<strong>特定の用途向けに設計されたサービス</strong>の1つにすぎません。"
        f"試験ではサービス名より、背後にある<strong>LLM・エージェント・自動化</strong>などの概念が問われることもあります。"
    )

    faqs = [
        (f"{name}は無料で使えますか？", f"無料枠または試用がある場合もありますが、本格利用は有料プランが中心です。最新の料金は公式サイトで確認してください（2026年6月時点）。"),
        (f"{name}とChatGPTの違いは？", f"{name}は{maker}が提供する{cat_short}向けツールです。ChatGPTは汎用チャットAIで、用途と連携先が異なります。目的に応じて使い分けるのがおすすめです。"),
        (f"{name}は何社が提供していますか？", f"{maker}が提供するサービスです。試験では開発元とサービス名をセットで覚えると整理しやすくなります。"),
        (f"会社の業務で{name}を使っても大丈夫ですか？", "個人の無料アカウントに機密情報を入力するのは避けるべきです。業務利用では社内ガイドラインの整備と、出力内容の人による確認が推奨されます。"),
        (f"資格試験で{name}は出ますか？", f"生成AIパスポートやG検定では、{cat_short}の活用例や注意点として類似の設問が出ることがあります。サービス名そのものより、できること・リスクの整理が重要です。"),
    ]

    faq_html = "\n".join(faq_item(q, a) for q, a in faqs)
    faq_ld = ",\n          ".join(faq_json(q, a) for q, a in faqs)

    related_ids = COMPARE_WITH.get(tid, compare_ids)
    related_html = "\n".join(
        related_link(rid, all_names.get(rid, rid)) for rid in related_ids if rid in all_names
    )

    compare_rows = ""
    for i, cid in enumerate(compare_ids):
        label = "◎" if i == 0 else "○" if i == 1 else "△"
        compare_rows += f"          <tr><td>{html.escape(all_names.get(cid, cid))}</td><td>{label if i==0 else '○' if i==1 else '△'}</td><td>{'◎' if i==1 else '○' if i==0 else '△'}</td><td>{'◎' if i==2 else '○'}</td></tr>\n"

    meta_desc = f"{name}（{maker}）の機能・料金・使い方を解説。{short_tagline(reason, 60)} G検定・生成AIパスポート対策にも。"[:160]

    return f"""<!DOCTYPE html>
<html lang="ja">
<head>
  <meta charset="UTF-8">
  <meta name="viewport" content="width=device-width, initial-scale=1.0">
  <meta name="description" content="{html.escape(meta_desc)}">
  <title>{html.escape(page_title)}</title>
  <link rel="canonical" href="{canonical}">
  <link rel="icon" href="/assets/images/favicon.svg" type="image/svg+xml">
  <link rel="icon" href="/assets/images/favicon-32x32.png" type="image/png" sizes="32x32">
  <link rel="icon" href="/assets/images/favicon-16x16.png" type="image/png" sizes="16x16">
  <link rel="apple-touch-icon" href="/assets/images/apple-touch-icon.png">
  <link rel="manifest" href="/site.webmanifest">
  <meta name="theme-color" content="#1A5CDB">
  <meta property="og:type" content="article">
  <meta property="og:site_name" content="AI Master">
  <meta property="og:title" content="{html.escape(headline)}">
  <meta property="og:description" content="{html.escape(meta_desc[:120])}">
  <meta property="og:url" content="{canonical}">
  <meta property="og:locale" content="ja_JP">
  <meta property="og:image" content="{og_url}">
  <meta name="twitter:card" content="summary_large_image">
  <meta name="twitter:image" content="{og_url}">
  <link rel="preconnect" href="https://fonts.googleapis.com">
  <link rel="preconnect" href="https://fonts.gstatic.com" crossorigin>
  <link href="https://fonts.googleapis.com/css2?family=Noto+Sans+JP:wght@400;500;600;700&display=swap" rel="stylesheet">
  <link rel="stylesheet" href="../../assets/css/main.css">
  <link rel="stylesheet" href="../../assets/css/tools.css">
  <link rel="stylesheet" href="../../assets/css/seo.css">
  <link rel="stylesheet" href="../../assets/css/tool-detail.css">
  <script type="application/ld+json">
  {{
    "@context": "https://schema.org",
    "@graph": [
      {{
        "@type": "BreadcrumbList",
        "@id": "{canonical}#breadcrumb",
        "itemListElement": [
          {{ "@type": "ListItem", "position": 1, "name": "ホーム", "item": "https://ai-master.jp/" }},
          {{ "@type": "ListItem", "position": 2, "name": "AIツール", "item": "https://ai-master.jp/tools/" }},
          {{ "@type": "ListItem", "position": 3, "name": "{html.escape(name)}", "item": "{canonical}" }}
        ]
      }},
      {{
        "@type": "Article",
        "@id": "{canonical}#article",
        "headline": "{html.escape(headline)}",
        "description": "{html.escape(meta_desc)}",
        "url": "{canonical}",
        "datePublished": "2026-06-10",
        "dateModified": "2026-06-10",
        "inLanguage": "ja",
        "image": "{og_url}",
        "author": {{ "@type": "Organization", "name": "AI Master" }},
        "publisher": {{ "@type": "Organization", "name": "AI Master", "url": "https://ai-master.jp/" }},
        "isPartOf": {{ "@id": "https://ai-master.jp/#website" }},
        "mainEntityOfPage": {{ "@id": "{canonical}#webpage" }},
        "about": {{ "@type": "SoftwareApplication", "name": "{html.escape(name)}" }}
      }},
      {{
        "@type": "WebPage",
        "@id": "{canonical}#webpage",
        "url": "{canonical}",
        "name": "{html.escape(headline)}",
        "description": "{html.escape(meta_desc)}",
        "isPartOf": {{ "@id": "https://ai-master.jp/#website" }},
        "breadcrumb": {{ "@id": "{canonical}#breadcrumb" }},
        "inLanguage": "ja"
      }},
      {{
        "@type": "SoftwareApplication",
        "name": "{html.escape(name)}",
        "applicationCategory": "BusinessApplication",
        "operatingSystem": "Web",
        "description": "{html.escape(short_tagline(reason, 120))}",
        "provider": {{ "@type": "Organization", "name": "{html.escape(maker)}" }}
      }},
      {{
        "@type": "FAQPage",
        "mainEntity": [
          {faq_ld}
        ]
      }}
    ]
  }}
  </script>
  <!-- Google tag (gtag.js) -->
  <script async src="https://www.googletagmanager.com/gtag/js?id=G-FWXFGSH6TD"></script>
  <script>
    window.dataLayer = window.dataLayer || [];
    function gtag(){{dataLayer.push(arguments);}}
    gtag('js', new Date());
    gtag('config', 'G-FWXFGSH6TD');
  </script>
</head>
<body class="page-body">

<nav class="site-nav">
  <a href="../../index.html" class="logo">AI<em>マスター</em></a>
  <ul class="nav-links">
    <li><a href="../../exams/">試験対策</a></li>
    <li><a href="../../guide/">学習ガイド</a></li>
    <li><a href="../../glossary/">用語辞典</a></li>
    <li><a href="../index.html" class="active">AIツール</a></li>
    <li><a href="../../career/">キャリア</a></li>
  </ul>
</nav>

<main class="page-wrap page-wrap--article">
  <nav class="breadcrumb" aria-label="パンくずリスト">
    <ol>
      <li><a href="../../index.html">ホーム</a></li>
      <li><a href="../index.html">AIツール</a></li>
      <li aria-current="page">{html.escape(name)}</li>
    </ol>
  </nav>

  <header class="tool-hero tool-hero--split">
    <div class="tool-hero-body">
      <div class="tool-hero-brand">
        <img src="../../assets/images/tools/{tid}/{app_icon}" alt="" width="52" height="52" class="tool-hero-app-icon">
        <div>
          <p class="tool-hero-eyebrow">{html.escape(maker)} · {html.escape(cat_short)}</p>
          <h1>{html.escape(headline)}</h1>
        </div>
      </div>
      <p class="tool-hero-maker">{html.escape(tool["tagline"])}</p>
      <div class="tool-hero-meta">
        <span class="tool-category">無料枠あり</span>
        <span class="tool-category">有料プランあり</span>
      </div>
      <div class="tool-platform-bar">
        <span class="tool-platform-label">対応環境</span>
        <ul class="tool-platform-chips">
          <li>Web</li>
        </ul>
      </div>
    </div>
    <figure class="tool-hero-visual">
      <img src="../../assets/images/tools/{tid}/{og_hero}" alt="{html.escape(name)}の公式イメージ" width="800" height="450" fetchpriority="high">
      <figcaption>出典：{html.escape(maker)}公式</figcaption>
    </figure>
  </header>

  <div class="tool-meta-bar">
    <span>更新日：<time datetime="2026-06-10">2026年6月10日</time></span>
    <span>読了目安：約8分</span>
  </div>

  <article class="tool-content">
    <p class="tool-lead">
      <strong>{html.escape(name)}</strong>は、{html.escape(maker)}が提供する{html.escape(cat_short)}向けのAIツールです。本記事は機能の網羅リストではなく、<strong>G検定・生成AIパスポートで押さえる定義</strong>と<strong>実務で起きやすい誤解</strong>に絞って整理しています。料金・機能は<strong>2026年6月時点</strong>の情報です。契約前は<a href="#pricing">料金セクション</a>を参照し、必ず公式でも確認してください。
    </p>

    <nav class="tool-toc" aria-label="目次">
      <p class="tool-toc-title">目次</p>
      <ol>
        <li><a href="#exam-study">試験で問われる見方</a></li>
        <li><a href="#what-is">{html.escape(name)}とは</a></li>
        <li><a href="#features">できること（主な機能）</a></li>
        <li><a href="#misconception">よくある誤解</a></li>
        <li><a href="#pricing">料金（2026年6月時点）</a></li>
        <li><a href="#how-to">はじめ方・基本的な使い方</a></li>
        <li><a href="#use-cases">活用例</a></li>
        <li><a href="#pros-cons">メリット・デメリット</a></li>
        <li><a href="#comparison">主要ツールとの比較</a></li>
        <li><a href="#recommend">こんな人におすすめ</a></li>
        <li><a href="#faq">よくある質問</a></li>
      </ol>
    </nav>

    <section id="exam-study">
      <h2>試験で問われる見方</h2>
      <p>生成AIパスポートでは、{html.escape(cat_short)}の業務活用（できること・注意点・情報の取り扱い）が論点になります。{html.escape(name)}は「{html.escape(short_tagline(reason, 40))}」として整理すると覚えやすいです。</p>
      <p>{html.escape(reason[:200])}{"…" if len(reason) > 200 else ""}</p>
      <div class="tool-callout">
        <p><strong>このサイトの演習で確認する</strong></p>
        <p>生成AIパスポート：<a href="../../exams/genai-passport/drill/q/tf-0201/">一問一答 TF-0201（テキスト生成AIの活用）</a> · <a href="../../exams/genai-passport/drill/">一問一答一覧</a></p>
        <p>G検定：<a href="../../exams/g-kentei/drill/q/tf-170/">一問一答 TF-170（生成AI）</a> · <a href="../../exams/g-kentei/drill/">一問一答一覧</a></p>
      </div>
    </section>

    <section id="what-is">
      <h2>{html.escape(name)}とは</h2>
      <p>{html.escape(name)}は、{html.escape(maker)}が提供する{html.escape(cat_short)}向けサービスです。{html.escape(reason.split("。")[0])}。</p>
      <p>単なるチャットボットではなく、<strong>特定のワークフローに組み込まれたAI</strong>として設計されている点が特徴です。試験ではサービス名と開発元、できることの範囲をセットで覚えると得点源になります。</p>
    </section>

    <section id="features">
      <h2>できること（主な機能）</h2>
      <div class="tool-feature-grid">
        <div class="tool-feature-card"><h3>AI生成・支援</h3><p>テキスト・資料・メディアなど、{html.escape(cat_short)}向けの出力をAIが支援します。</p></div>
        <div class="tool-feature-card"><h3>テンプレート・ワークフロー</h3><p>繰り返し作業をテンプレート化し、品質を一定に保ちやすくします。</p></div>
        <div class="tool-feature-card"><h3>連携・エクスポート</h3><p>他ツールとの連携や、成果物の書き出しに対応している場合があります（公式仕様を確認）。</p></div>
        <div class="tool-feature-card"><h3>チーム利用</h3><p>法人・チーム向けプランでは権限管理や監査ログが提供される場合があります。</p></div>
        <div class="tool-feature-card"><h3>カスタマイズ</h3><p>プロンプト・設定・ブランドガイドに沿った出力調整が可能な場合があります。</p></div>
        <div class="tool-feature-card"><h3>API連携</h3><p>開発者向けにAPIやWebhookが用意されている場合、業務システムと統合できます。</p></div>
      </div>
    </section>

    <section id="misconception">
      <h2>よくある誤解</h2>
      <p>{misconception}</p>
      <p>もう1つは「<strong>有料プランなら出力が常に正確</strong>」という期待です。どのプランでもハルシネーションや古い情報の混入は起こり得るため、重要な判断は人が確認する必要があります。</p>
    </section>

    <section id="pricing">
      <h2>料金（2026年6月時点）</h2>
      <p>{html.escape(name)}の料金体系は改定されやすいため、以下は<strong>2026年6月時点</strong>の整理です。導入・契約前に必ず公式のPricingページで最新情報を確認してください。</p>
      <table class="tool-spec-table">
        <thead><tr><th>プラン</th><th>目安</th><th>向いている人</th></tr></thead>
        <tbody>
          <tr><td><strong>無料 / 試用</strong></td><td>$0</td><td>まず触ってみたい個人</td></tr>
          <tr><td><strong>個人有料</strong></td><td>月額課金</td><td>継続利用するフリーランス</td></tr>
          <tr><td><strong>チーム / 法人</strong></td><td>座席課金</td><td>社内展開・権限管理が必要な組織</td></tr>
        </tbody>
      </table>
    </section>

    <section id="how-to">
      <h2>はじめ方・基本的な使い方</h2>
      <ol class="tool-steps">
        <li><strong>アカウント作成</strong>公式サイトからサインアップし、利用規約とデータの取り扱いを確認します。</li>
        <li><strong>目的を決める</strong>試したいユースケース（例：資料作成・自動化・生成）を1つに絞ります。</li>
        <li><strong>テンプレートを選ぶ</strong>用意されているテンプレートやサンプルから始めると学習コストが下がります。</li>
        <li><strong>出力を検証</strong>事実関係・数値・固有名詞は必ず人が確認します。</li>
        <li><strong>業務利用の可否を確認</strong>機密データを入力する前に、社内ガイドラインとプランのデータポリシーを確認します。</li>
      </ol>
    </section>

    <section id="use-cases">
      <h2>活用例</h2>
      <ul>
        <li>{html.escape(cat_short)}の下書き・反復作業の効率化</li>
        <li>チーム内のナレッジ整理と共有</li>
        <li>資格試験の学習で「具体サービスの例」として押さえる</li>
        <li>既存ツール（メール・チャット・IDE等）との連携による業務短縮</li>
      </ul>
    </section>

    <section id="pros-cons">
      <h2>メリット・デメリット</h2>
      <table class="tool-pros-cons-table">
        <thead><tr><th>メリット</th><th>デメリット</th></tr></thead>
        <tbody>
          <tr><td>{html.escape(cat_short)}に特化したワークフロー</td><td>汎用チャットAIほど柔軟ではない場合がある</td></tr>
          <tr><td>作業時間の短縮</td><td>出力の事実確認が依然必要</td></tr>
          <tr><td>チーム展開しやすい</td><td>料金・上限が用途によっては高くなり得る</td></tr>
          <tr><td>試験の具体例として覚えやすい</td><td>機能改定が速く記事が古くなりやすい</td></tr>
        </tbody>
      </table>
    </section>

    <section id="comparison">
      <h2>{html.escape(compare_names[0])}などとの比較</h2>
      <table class="tool-spec-table">
        <thead><tr><th>観点</th><th>{html.escape(name)}</th><th>{html.escape(compare_names[0])}</th><th>{html.escape(compare_names[1] if len(compare_names)>1 else "汎用AI")}</th></tr></thead>
        <tbody>
          <tr><td>特化領域</td><td>{html.escape(cat_short)}</td><td>同カテゴリ</td><td>汎用</td></tr>
          <tr><td>開発元</td><td>{html.escape(maker)}</td><td>—</td><td>—</td></tr>
          <tr><td>試験での覚え方</td><td>サービス名＋用途</td><td>比較対象として整理</td><td>概念の理解</td></tr>
          <tr><td>向いている人</td><td>{html.escape(cat_short)}を深く使う人</td><td>代替候補を比較したい人</td><td>まず1本試す人</td></tr>
        </tbody>
      </table>
    </section>

    <section id="recommend">
      <h2>こんな人におすすめ</h2>
      <ul>
        <li><strong>{html.escape(cat_short)}を本業で使う担当者</strong></li>
        <li><strong>資格試験で具体サービスの例が欲しい受験生</strong></li>
        <li><strong>{html.escape(compare_names[0])}などと比較検討している人</strong></li>
      </ul>
    </section>

    <section id="faq">
      <h2>よくある質問</h2>
      <div class="tool-faq">
{faq_html}
      </div>
    </section>

    <p class="tool-product-links tool-product-links--footer">
      公式リンク：
      <a href="{html.escape(url)}" target="_blank" rel="noopener noreferrer">{html.escape(name)}</a>
    </p>

    <aside class="tool-related-wrap">
      <div class="tool-related-block">
        <h2 class="tool-related-heading">関連するAIツール</h2>
        <ul class="tool-related-list">
{related_html}
        </ul>
      </div>
    </aside>
  </article>
</main>

<footer>
  <div class="foot-inner">
    <div class="foot-top">
      <div>
        <div class="foot-logo">AI<em>マスター</em></div>
        <p class="foot-tagline">AIスキルを、キャリアの武器に。</p>
      </div>
      <div class="foot-cols">
        <div class="foot-col">
          <p class="foot-col-h">学ぶ</p>
          <a href="../../glossary/">AI用語辞典</a>
          <a href="../index.html">AIツール</a>
        </div>
        <div class="foot-col">
          <p class="foot-col-h">試験対策</p>
          <a href="../../exams/g-kentei/">G検定対策</a>
          <a href="../../exams/genai-passport/">生成AIパスポート対策</a>
          <a href="../../exams/">資格一覧</a>
        </div>
        <div class="foot-col">
          <p class="foot-col-h">キャリア</p>
          <a href="../../career/">AI職種ガイド</a>
        </div>
      </div>
    </div>
    <p class="foot-disclaimer">G検定は（社）日本ディープラーニング協会の登録商標です。生成AIパスポートは一般社団法人グロービスの商標です。当サイトは各試験の公式サイトではありません。</p>
    <div class="foot-bottom">
      <span>© 2026 AIマスター. All rights reserved.</span>
      <span class="foot-legal">
        <a href="../../legal/privacy.html">プライバシーポリシー</a>
        ·
        <a href="../../legal/terms.html">利用規約</a>
      </span>
    </div>
  </div>
</footer>

<script src="../../assets/js/site-config.js"></script>
<script src="../../assets/js/seo.js"></script>
</body>
</html>
"""


def update_sitemap(tool_ids: list[str]) -> None:
    text = SITEMAP.read_text(encoding="utf-8")
    added = 0
    for tid in tool_ids:
        loc = f"https://ai-master.jp/tools/{tid}/"
        if loc in text:
            continue
        block = (
            f"  <url>\n"
            f"    <loc>{loc}</loc>\n"
            f"    <changefreq>monthly</changefreq>\n"
            f"    <priority>0.8</priority>\n"
            f"  </url>\n"
        )
        text = text.replace(
            "  <url>\n    <loc>https://ai-master.jp/tools/whisper/</loc>",
            block + "  <url>\n    <loc>https://ai-master.jp/tools/whisper/</loc>",
            1,
        )
        added += 1
    SITEMAP.write_text(text, encoding="utf-8")
    print(f"sitemap: {added} URL 追加")


def main() -> None:
    if not CANDIDATES_XLSX.is_file():
        raise SystemExit(f"候補ファイルがありません: {CANDIDATES_XLSX}")

    candidates = load_candidates()
    merge_excel(candidates)

    # build_tools_data must run after excel merge — import subprocess
    import subprocess

    subprocess.run(
        ["python3", str(ROOT / "tools" / "build_tools_data.py")],
        check=True,
        cwd=str(ROOT),
    )

    all_names = load_all_tool_names()
    # reload after build
    all_names = load_all_tool_names()
    for c in candidates:
        all_names[c["id"]] = c["name"]

    written = 0
    new_ids = []
    for tool in candidates:
        out_dir = ROOT / "tools" / tool["id"]
        out_dir.mkdir(parents=True, exist_ok=True)
        out_path = out_dir / "index.html"
        if out_path.is_file():
            continue
        out_path.write_text(generate_article(tool, all_names), encoding="utf-8")
        written += 1
        new_ids.append(tool["id"])
        print(f"  記事: {tool['id']}")

    subprocess.run(["python3", str(ROOT / "tools" / "build_tools_data.py")], check=True, cwd=str(ROOT))
    subprocess.run(["python3", str(ROOT / "tools" / "build_hub_static_lists.py")], check=True, cwd=str(ROOT))
    if new_ids:
        update_sitemap(new_ids)

    print(f"\n完了: 新規記事 {written} 本 / 候補 {len(candidates)} 件")


if __name__ == "__main__":
    main()
