#!/usr/bin/env python3
"""AIモデル一覧v2.xlsx から用語辞典マスタ（glossary-terms.csv）へモデル語を追加する。"""

from __future__ import annotations

import argparse
import csv
import re
import shutil
import sys
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT / "tools"))

from glossary_terms import CSV_FIELDS, CSV_FILE, load_terms_csv, validate_terms, write_csv  # noqa: E402

import openpyxl

XLSX_DEFAULT = Path.home() / "Downloads" / "AIモデル一覧v2.xlsx"
XLSX_PROJECT = ROOT / "data" / "ai-models-v2.xlsx"
CATEGORY = "models-tech"
DEFAULT_EXAMS = "g-kentei"
DEFAULT_TIER = "2"
DEFAULT_STATUS = "planned"

SLUG_RE = re.compile(r"^[a-z0-9]+(?:-[a-z0-9]+)*$")


def slugify(name: str) -> str:
    base = re.sub(r"（[^）]*）", "", str(name)).strip()
    base = base.replace("·", "-").replace(".", "-")
    slug = re.sub(r"[^a-zA-Z0-9]+", "-", base).strip("-").lower()
    aliases = {
        "gpt-3-5": "gpt-3-5",
        "gpt-4-turbo": "gpt-4-turbo",
        "gpt-4o-mini": "gpt-4o-mini",
        "o1-mini": "o1-mini",
        "o3-mini": "o3-mini",
        "o4-mini": "o4-mini",
        "command-a": "command-a",
        "gemma-2": "gemma-2",
        "qwen2-5": "qwen2-5",
        "flux-1": "flux-1",
    }
    return aliases.get(slug, slug)


def clean_name(name: str) -> str:
    return re.sub(r"（[^）]*）", "", str(name)).strip()


def parse_summary(note: str) -> str:
    text = str(note).strip()
    if "。" in text:
        tail = text.split("。", 1)[1].strip()
        if tail:
            return tail if tail.endswith("。") else f"{tail}。"
    return text


def reading_key(name: str) -> str:
    for ch in name:
        if ch.isalnum():
            return ch.upper()
    return name[:1] or "?"


def resolve_maker(maker: str, name: str) -> str:
    maker = str(maker).strip()
    if maker and maker != "その他":
        return maker
    paren = re.search(r"（([^）]+)）", str(name))
    return paren.group(1).strip() if paren else maker or "その他"


def load_models(xlsx: Path) -> list[dict]:
    wb = openpyxl.load_workbook(xlsx, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows: list[dict] = []
    for r in range(2, ws.max_row + 1):
        maker = ws.cell(r, 1).value
        name = ws.cell(r, 2).value
        note = ws.cell(r, 3).value
        if not name or not note:
            continue
        display_name = clean_name(str(name))
        rows.append(
            {
                "maker": resolve_maker(str(maker or ""), str(name)),
                "name": display_name,
                "summary": parse_summary(str(note)),
                "id": slugify(display_name),
            }
        )
    return rows


def existing_keys(terms: list[dict]) -> tuple[set[str], set[str]]:
    ids = {t["id"] for t in terms}
    names = {t["name"].lower() for t in terms}
    return ids, names


def should_skip(model: dict, ids: set[str], names: set[str]) -> str | None:
    sid = model["id"]
    name = model["name"].lower()
    if sid in ids:
        return f"id '{sid}' が既存"
    if name in names:
        return f"name '{model['name']}' が既存"
    if f"{sid}-model" in ids:
        return f"ファミリー語 '{sid}-model' が既存"
    return None


def model_to_csv_row(model: dict) -> dict:
    yomi = model["name"]
    return {
        "id": model["id"],
        "name": model["name"],
        "yomi": yomi,
        "reading": reading_key(model["name"]),
        "summary": model["summary"],
        "category": CATEGORY,
        "sort_key": reading_key(model["name"]),
        "status": DEFAULT_STATUS,
        "tier": DEFAULT_TIER,
        "exams": DEFAULT_EXAMS,
        "priority": "",
        "notes": f"開発元: {model['maker']}",
    }


def main() -> None:
    parser = argparse.ArgumentParser(description="AIモデル一覧 Excel を用語辞典 CSV に取り込む")
    parser.add_argument("--dry-run", action="store_true", help="追加せず件数のみ表示")
    args = parser.parse_args()

    src = XLSX_DEFAULT if XLSX_DEFAULT.exists() else XLSX_PROJECT
    if not src.exists():
        raise SystemExit(f"Excel not found: {src}")

    XLSX_PROJECT.parent.mkdir(parents=True, exist_ok=True)
    if src != XLSX_PROJECT:
        shutil.copy2(src, XLSX_PROJECT)

    models = load_models(src)
    terms = load_terms_csv()
    ids, names = existing_keys(terms)

    to_add: list[dict] = []
    skipped: list[tuple[str, str]] = []
    for model in models:
        reason = should_skip(model, ids, names)
        if reason:
            skipped.append((model["name"], reason))
            continue
        if model["id"] in ids or model["name"].lower() in names:
            skipped.append((model["name"], "重複"))
            continue
        row = model_to_csv_row(model)
        to_add.append(row)
        ids.add(row["id"])
        names.add(row["name"].lower())

    print(f"Excel: {len(models)} 件")
    print(f"追加予定: {len(to_add)} 件")
    print(f"スキップ: {len(skipped)} 件")
    for name, reason in skipped:
        print(f"  - {name}: {reason}")

    if args.dry_run:
        return

    if not to_add:
        print("追加する用語はありません。")
        return

    merged = terms + to_add
    errors = validate_terms(merged)
    if errors:
        raise SystemExit("検証エラー:\n" + "\n".join(errors[:20]))

    write_csv(merged)
    print(f"updated {CSV_FILE} (+{len(to_add)} → 合計 {len(merged)} 語)")


if __name__ == "__main__":
    main()
